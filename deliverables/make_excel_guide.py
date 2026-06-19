"""
make_excel_guide.py
Builds a comprehensive Excel User Guide for the Bank Assistant Capstone demo.
Sheets:
  1. Cover               — title, contents, how-to-present
  2. Scenario Overview   — all 8 scenarios in a table with descriptions
  3-10. One sheet per scenario — screenshot + full comm flow table
  11. Agent Communication Map  — who calls whom
  12. Skill-Hook Reference     — all skills and hooks in one view
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_TEXT
import os

SCREENSHOTS = "/home/labuser/Day10/deliverables/screenshots"
OUT = "/home/labuser/Day10/deliverables/06_Presentation_User_Guide.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "00336699"  # header backgrounds
MID_BLUE    = "004472C4"
LIGHT_BLUE  = "FFDCE6F1"
PALE_BLUE   = "FFF2F7FF"
GREEN       = "FF00823E"
LIGHT_GREEN = "FFE2EFDA"
RED         = "FFC00000"
LIGHT_RED   = "FFFCE4D6"
ORANGE      = "FFB44A00"
LIGHT_ORANGE= "FFFFEB9C"
PURPLE      = "FF6400A0"
LIGHT_PURPLE= "FFEDDEF7"
TEAL        = "FF006464"
LIGHT_TEAL  = "FFD6EAE8"
GRAY        = "FF404040"
LIGHT_GRAY  = "FFF2F2F2"
WHITE       = "FFFFFFFF"
GOLD        = "FFFFC107"
DARK_TEXT   = "FF1A1A2E"
YELLOW_BG   = "FFFFFF99"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(name="Calibri", size=11, bold=False, italic=False, color="FF000000", underline=False):
    return Font(name=name, size=size, bold=bold, italic=italic,
                color=color, underline="single" if underline else None)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="FF004472")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, fnt=None, fll=None, aln=None, bdr=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fll: c.fill = fll
    if aln: c.alignment = aln
    if bdr: c.border = bdr
    return c

def header_row(ws, row, headers, widths, fill_color=DARK_BLUE, text_color=WHITE, size=11):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = set_cell(ws, row, i, h,
                     fnt=font(size=size, bold=True, color=text_color),
                     fll=fill(fill_color),
                     aln=align("center", "center"),
                     bdr=thin_border())
        ws.column_dimensions[get_column_letter(i)].width = w

def data_row(ws, row, values, fll_hex=WHITE, text_color="FF1A1A2E", size=10,
             bold_col=None, wrap_cols=None):
    for i, v in enumerate(values, start=1):
        b = (bold_col is not None and i == bold_col)
        wt = (wrap_cols is not None and i in wrap_cols)
        set_cell(ws, row, i, v,
                 fnt=font(size=size, bold=b, color=text_color),
                 fll=fill(fll_hex if (row % 2 == 0) else WHITE),
                 aln=align("left", "top", wrap=wt),
                 bdr=thin_border())

def merge_title(ws, row, col_start, col_end, text, fll_hex=DARK_BLUE,
                txt_color=WHITE, size=13, height=28):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = font(size=size, bold=True, color=txt_color)
    c.fill = fill(fll_hex)
    c.alignment = align("center", "center")
    ws.row_dimensions[row].height = height

def section_label(ws, row, col_start, col_end, text, fll_hex=MID_BLUE, size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = font(size=size, bold=True, color=WHITE)
    c.fill = fill(fll_hex)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 22


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: COVER
# ════════════════════════════════════════════════════════════════════════════
def sheet_cover():
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 3

    ws.row_dimensions[1].height = 20
    merge_title(ws, 2, 2, 3, "🏦  NexaBank Internal Bank Employee Assistant", DARK_BLUE, WHITE, 20, 50)
    merge_title(ws, 3, 2, 3, "Presentation User Guide — Demo Scenarios & Agent Communication Map",
                MID_BLUE, WHITE, 13, 32)
    merge_title(ws, 4, 2, 3, "Capstone Project  |  Multi-Agent AI System  |  June 2025",
                "FF2E4057", "FFFFC107", 11, 24)

    ws.row_dimensions[5].height = 14
    # Contents table
    section_label(ws, 6, 2, 3, "  📋  Workbook Contents", MID_BLUE)
    contents = [
        ("Cover",                   "This page — overview and how to use this guide"),
        ("Scenario Overview",       "All 8 demo scenarios at a glance (quick reference table)"),
        ("SC1 — Loan Policy",       "PolicySearchAgent: keyword match → single policy answer"),
        ("SC2 — KYC Policy",        "PolicySearchAgent: KYC domain keyword match"),
        ("SC3 — Multi-Policy",      "MultiPolicyAgent: cross-domain query (KYC + Loan)"),
        ("SC4 — Summary",           "SummaryAgent: 5-bullet policy overview on demand"),
        ("SC5 — Escalation",        "EscalationAgent: contact routing + email draft"),
        ("SC6 — Compliance Alert",  "compliance_alert_hook: risk keyword detection in real time"),
        ("SC7 — Fallback Self-Heal","FallbackAgent: self-healing when no policy matches"),
        ("SC8 — Account Opening",   "PolicySearchAgent: account domain keyword match"),
        ("Agent Communication Map", "Full diagram of which agent calls which skill/hook"),
        ("Skill & Hook Reference",  "All 3 skills and 6 hooks: inputs, outputs, trigger conditions"),
    ]
    header_row(ws, 7, ["Sheet", "What it shows"], [30, 55], MID_BLUE)
    for r, (sheet, desc) in enumerate(contents, start=8):
        alt = LIGHT_BLUE if r % 2 == 0 else WHITE
        data_row(ws, r, [sheet, desc], fll_hex=alt[2:] if alt.startswith("FF") else alt,
                 bold_col=1, wrap_cols={2})
        ws.row_dimensions[r].height = 20

    ws.row_dimensions[r+1].height = 14
    section_label(ws, r+2, 2, 3, "  🎯  How to Use This Guide During Your Presentation", "FF00823E")
    tips = [
        ("Step 1",  "Open this guide alongside the live app in a second window or printed copy"),
        ("Step 2",  "Navigate to each Scenario sheet to see the EXACT question to type in the app"),
        ("Step 3",  "The 'User Input' row shows the question — type it into the chatbot exactly"),
        ("Step 4",  "The communication flow table shows which agent fires and why — explain this to your audience"),
        ("Step 5",  "Point to the screenshot to show what the employee sees in the UI"),
        ("Step 6",  "Use the 'Talking Points' column to explain the technical concept behind each step"),
        ("Tip",     "For SC3 (Multi-Policy), explain the cross_reference scoring threshold (top ≥ 2 AND second ≥ 1)"),
        ("Tip",     "For SC6 (Compliance), show how the yellow banner appears instantly — compliance_alert_hook fires before the UI renders"),
        ("Tip",     "For SC7 (Fallback), emphasize that the employee never saw the PolicySearchAgent fail — self-healing is invisible"),
    ]
    header_row(ws, r+3, ["", "Guidance"], [8, 77], "FF00823E")
    for ri, (step, tip) in enumerate(tips, start=r+4):
        alt = LIGHT_GREEN if ri % 2 == 0 else WHITE
        data_row(ws, ri, [step, tip], fll_hex=alt[2:] if alt != WHITE else WHITE,
                 bold_col=1, wrap_cols={2})
        ws.row_dimensions[ri].height = 20

    return ws


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: SCENARIO OVERVIEW
# ════════════════════════════════════════════════════════════════════════════
def sheet_scenario_overview():
    ws = wb.create_sheet("Scenario Overview")
    ws.sheet_view.showGridLines = False

    widths = [4, 22, 28, 20, 18, 22, 20]
    cols = ["#", "Scenario Name", "Question to Type", "Intent Classified",
            "Agent Invoked", "Key Feature Demonstrated", "Expected Result"]

    section_label(ws, 1, 1, 7, "  🏦  NexaBank Bank Assistant — All 8 Presentation Scenarios", DARK_BLUE, 13)
    header_row(ws, 2, cols, widths, MID_BLUE)

    scenarios = [
        ("SC1", "Loan Policy Search",
         "What documents are needed for a loan?",
         "policy_query", "PolicySearchAgent",
         "Keyword scoring on KEYWORD_MAP; PolicySearchAgent returns best snippet",
         "Loan policy answer + LP-001 badge + PolicySearchAgent chip"),

        ("SC2", "KYC Policy Search",
         "What are the KYC requirements for a new customer?",
         "policy_query", "PolicySearchAgent",
         "KYC keyword cluster match; extract_answer() extracts verbatim passage",
         "KYC answer + KYC-002 badge displayed"),

        ("SC3", "Multi-Policy Query",
         "What KYC documents do I need for a loan application?",
         "policy_query (multi)", "MultiPolicyAgent",
         "cross_reference() scores both KYC (≥2) and Loan (≥1) — threshold met",
         "Two source badges; merged answer from both policies with divider"),

        ("SC4", "Policy Summary",
         "Summarize the credit card policy",
         "summary", "SummaryAgent",
         "SummaryAgent uses POLICY_NAME_HINTS ('card') to detect policy; summarize_policy() produces 5 bullets",
         "5-bullet summary + CCP-004 badge + SummaryAgent chip"),

        ("SC5", "Escalation Request",
         "I need to escalate a KYC dispute to the compliance team",
         "escalation", "EscalationAgent",
         "EscalationAgent reads policy files to infer team; generates email draft via generate_escalation_email()",
         "Contact details + email draft expander + escalated=True in result"),

        ("SC6", "Compliance Alert",
         "How do I bypass KYC with a fake document?",
         "policy_query", "PolicySearchAgent + compliance_alert_hook",
         "compliance_alert_hook fires before UI renders; detect_compliance_risk() matched 'bypass' + 'fake document'",
         "Yellow warning banner at top of answer + COMPLIANCE_ALERT in audit log"),

        ("SC7", "Fallback / Self-Healing",
         "What is the HR leave policy for annual leave?",
         "policy_query → fallback", "FallbackAgent (auto-invoked)",
         "PolicySearchAgent scores 0 across all 5 policies; Orchestrator self-heals by chaining FallbackAgent",
         "Amber warning box; escalation_hook fires; employee sees polite no-match message"),

        ("SC8", "Account Opening Policy",
         "What is the minimum balance for a savings account?",
         "policy_query", "PolicySearchAgent",
         "'account' + 'savings' + 'balance' match account_opening policy keywords in KEYWORD_MAP",
         "Account policy answer + AOP-005 badge"),
    ]

    row_fills = [
        LIGHT_BLUE[2:], "E2EFDA", LIGHT_PURPLE[2:], LIGHT_BLUE[2:],
        LIGHT_RED[2:], LIGHT_ORANGE[2:], LIGHT_ORANGE[2:], "E2EFDA",
    ]
    for i, (sc, name, q, intent, agent, feature, result) in enumerate(scenarios, start=3):
        r = i
        fll = row_fills[i-3]
        ws.row_dimensions[r].height = 58
        for ci, val in enumerate([sc, name, q, intent, agent, feature, result], start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill(fll)
            c.border = thin_border()
            c.alignment = align("left", "top", wrap=True)
            c.font = font(size=10, bold=(ci <= 2))

    # column widths
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ════════════════════════════════════════════════════════════════════════════
# SCENARIO DETAIL SHEETS (SC1–SC8)
# ════════════════════════════════════════════════════════════════════════════

SCENARIO_DATA = {
    "SC1 — Loan Policy": {
        "tab_color": "004472C4",
        "screenshot": "sc1_loan_policy_search.png",
        "question": "What documents are needed for a loan?",
        "intent": "policy_query",
        "steps": [
            # (step, component, type, what_happens, talking_point, fill)
            ("1", "Employee", "UI Input",
             "Types question into chat input box",
             "Employee interacts only via the chat UI — completely abstracted from the agent system",
             "DCE6F1"),
            ("2", "app.py", "Streamlit",
             "Captures question text; retrieves OrchestratorAgent from @st.cache_resource; calls orchestrator.answer(question)",
             "OrchestratorAgent is cached — it's the same instance across all queries in one session",
             "EBF3FB"),
            ("3", "pre_query_hook", "Hook",
             "Validates: question is not empty, length > 3 chars. Returns {valid: True, question: sanitized}",
             "First gate — rejects empty or gibberish before any agent runs. Logs incoming query to session_events.log",
             "E2EFDA"),
            ("4", "QueryClassifierAgent", "SubAgent 1",
             "Calls extract_keywords(question) → ['documents', 'loan']. 'loan' not in summary/escalation triggers → intent = policy_query",
             "extract_keywords() filters stopwords and words < 4 chars. 'loan' survives, classifying as policy_query",
             "EBF3FB"),
            ("5", "OrchestratorAgent", "Orchestrator",
             "Calls cross_reference(question, policies) → scores: loan=2, others=0. top=2, second=0. Threshold not met → single domain. Routes to PolicySearchAgent",
             "Cross-reference threshold: top ≥ 2 AND second ≥ 1. Since second=0, single-policy path taken",
             "DCE6F1"),
            ("6", "search_skill.search_policy()", "Skill",
             "Scores each of 5 policies via KEYWORD_MAP. 'loan' matches loan_policy keywords (score=2). Returns {matched_key: loan_policy, score: 2, snippet: ...}",
             "KEYWORD_MAP is the pure keyword scorer — no ML, no LLM. The policy with the highest keyword count wins",
             "E2EFDA"),
            ("7", "PolicySearchAgent", "SubAgent 2",
             "Receives snippet from search_policy(). Calls format_skill.format_answer(snippet, 'loan_policy'). Returns {answer: '...', policy_key: 'loan_policy', found: True, score: 2}",
             "PolicySearchAgent wraps the skill calls — it never does text search itself, it orchestrates skills",
             "EBF3FB"),
            ("8", "AuditLoggerAgent", "SubAgent 7",
             "Receives (question, 'loan_policy', found=True). Calls audit_skill.log_query() → writes to query_audit.log. detect_compliance_risk() returns False",
             "Runs silently after every query regardless of which subagent answered. Employee never sees this",
             "DCE6F1"),
            ("9", "compliance_alert_hook", "Hook",
             "AuditLoggerAgent compliance_risk=False → hook NOT fired. No alert added to result",
             "The hook only fires when detect_compliance_risk() returns True — 14 risk keywords checked",
             "E2EFDA"),
            ("10", "escalation_hook", "Hook",
             "found=True → hook condition (found=False AND escalated=False) NOT met. Hook suppressed",
             "escalation_hook ONLY fires when the system has no answer AND no escalation was done",
             "EBF3FB"),
            ("11", "post_answer_hook", "Hook",
             "Logs: question, policy=Loan Policy LP-001, status=ANSWERED, response_ms to session_events.log",
             "Creates a full audit trail entry with timing for every successful answer",
             "DCE6F1"),
            ("12", "app.py UI render", "Streamlit",
             "Displays: agent chip (PolicySearchAgent | policy_query), policy badge (📚 Loan Policy LP-001), formatted answer in chat bubble",
             "The 10-field result dict is unpacked — policy_name drives the badge, agent_used drives the chip",
             "E2EFDA"),
        ],
        "result": "Loan Policy answer displayed | LP-001 badge | PolicySearchAgent chip | Audit logged",
    },

    "SC2 — KYC Policy": {
        "tab_color": "006400A0",
        "screenshot": "sc2_kyc_search.png",
        "question": "What are the KYC requirements for a new customer?",
        "intent": "policy_query",
        "steps": [
            ("1","Employee","UI Input","Types KYC question","Same chat UI — no change from employee perspective","DCE6F1"),
            ("2","app.py","Streamlit","Captures question; calls orchestrator.answer(question)","Single entry point — OrchestratorAgent from cache","EBF3FB"),
            ("3","pre_query_hook","Hook","Validates input — length > 3, not empty. Returns valid=True","Logs incoming to session_events.log","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","extract_keywords() → ['kyc', 'requirements', 'customer']. None match summary/escalation → policy_query","'kyc' and 'requirements' both > 3 chars, pass keyword filter","EBF3FB"),
            ("5","OrchestratorAgent","Orchestrator","cross_reference scores: kyc=3, loan=0, others=0. Second=0 → single-domain. Routes to PolicySearchAgent","Strong KYC signal; no ambiguity","DCE6F1"),
            ("6","search_skill.search_policy()","Skill","KEYWORD_MAP kyc_policy matches 'kyc', 'requirements', 'customer'. Returns snippet from kyc_policy.txt","extract_answer() pulls the most relevant paragraph from the policy file","E2EFDA"),
            ("7","PolicySearchAgent","SubAgent 2","format_answer() wraps snippet. Returns {found:True, policy_key:'kyc_policy', score:3}","format_answer adds header, cleans whitespace, adds reference code","EBF3FB"),
            ("8","AuditLoggerAgent","SubAgent 7","Logs to query_audit.log. No risk keywords → compliance_risk=False","Every query logged regardless of score","DCE6F1"),
            ("9","post_answer_hook","Hook","Logs ANSWERED status with response time","Full audit trail for every successful response","E2EFDA"),
            ("10","app.py UI","Streamlit","KYC-002 badge + PolicySearchAgent chip + answer displayed","Employee sees clean, formatted answer from the policy file","EBF3FB"),
        ],
        "result": "KYC policy answer | KYC-002 badge | Audit logged",
    },

    "SC3 — Multi-Policy": {
        "tab_color": "006400A0",
        "screenshot": "sc3_multi_policy.png",
        "question": "What KYC documents do I need for a loan application?",
        "intent": "policy_query (multi-domain)",
        "steps": [
            ("1","Employee","UI Input","Types cross-domain question referencing both KYC and loan","This is the test case that validates the MultiPolicyAgent threshold logic","DCE6F1"),
            ("2","app.py","Streamlit","Calls orchestrator.answer(question)","Same entry point","EBF3FB"),
            ("3","pre_query_hook","Hook","Validates input — passes","Sanitizes and logs","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","extract_keywords() → ['kyc','documents','loan','application']. All match policy_query intent","No summary or escalation triggers found — pure policy query","EBF3FB"),
            ("5","OrchestratorAgent routing","Orchestrator","Calls cross_reference(). KYC score=2, Loan score=1. Threshold check: top(2) ≥ 2 AND second(1) ≥ 1 → TRUE. Routes to MultiPolicyAgent","This is the key routing decision. If second score were 0, it would go to PolicySearchAgent instead","DCE6F1"),
            ("6","search_skill.cross_reference()","Skill","Returns sorted list: [{kyc_policy, score:2, snippet:...}, {loan_policy, score:1, snippet:...}]. Both above threshold","cross_reference() scores ALL 5 policies and returns all matches sorted desc","E2EFDA"),
            ("7","MultiPolicyAgent","SubAgent 3","Takes top 2 results. Fetches snippets from KYC policy and Loan policy. Merges with '─────' divider. Sets multi=True","The merge strategy: top snippet first, divider, second snippet — employee sees both policies in one answer","EBF3FB"),
            ("8","format_skill.format_answer()","Skill","Called twice — once per policy. Returns cleaned snippets for each","format_answer called by MultiPolicyAgent for each matched policy key","DCE6F1"),
            ("9","AuditLoggerAgent","SubAgent 7","Logs both matched policies. compliance_risk=False","Logs use multi flag; both policy names recorded","E2EFDA"),
            ("10","app.py UI","Streamlit","TWO source badges shown: KYC-002 (blue) + LP-001 (green). MultiPolicyAgent chip. Merged answer","Dual-badge UI distinguishes multi-policy answers visually from single-policy","EBF3FB"),
        ],
        "result": "Two policy answer merged | KYC-002 + LP-001 dual badges | MultiPolicyAgent chip",
    },

    "SC4 — Summary": {
        "tab_color": "006400A0",
        "screenshot": "sc4_summary.png",
        "question": "Summarize the credit card policy",
        "intent": "summary",
        "steps": [
            ("1","Employee","UI Input","Types summary request with keyword 'summarize'","The word 'summarize' is the trigger — any of: summarize, summary, overview, brief, explain the policy","DCE6F1"),
            ("2","app.py","Streamlit","Calls orchestrator.answer(question)","Standard pipeline entry","EBF3FB"),
            ("3","pre_query_hook","Hook","Valid input — passes","Logs and sanitizes","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","Detects 'summarize' in question → intent = summary. Returns {intent: 'summary'}","Summary triggers checked BEFORE general keyword scoring — intent takes priority","EBF3FB"),
            ("5","OrchestratorAgent","Orchestrator","Intent = summary → routes directly to SummaryAgent. No cross_reference needed for summary intent","Summary intent bypasses the multi-policy routing check entirely","DCE6F1"),
            ("6","SummaryAgent._detect_policy_key()","SubAgent 4 internal","Checks POLICY_NAME_HINTS: 'card' found in question → maps to credit_card_policy. Key resolved without calling search_policy()","POLICY_NAME_HINTS is faster than search_policy() — direct dict lookup. Only falls back to search if no hint matches","E2EFDA"),
            ("7","format_skill.summarize_policy()","Skill","Reads full credit_card_policy.txt. Extracts key sections. Returns 5-bullet string formatted summary","summarize_policy produces human-readable bullets — not raw policy text","EBF3FB"),
            ("8","SummaryAgent","SubAgent 4","Returns {answer: '5-bullet summary', policy_name: 'Credit Card Policy (CCP-004)', found: True}","SummaryAgent is the only agent that returns a curated summary rather than a verbatim extract","DCE6F1"),
            ("9","AuditLoggerAgent","SubAgent 7","Logs summary request. comply_risk=False for this question","Summary requests are audit-logged the same as search requests","E2EFDA"),
            ("10","app.py UI","Streamlit","SummaryAgent chip + CCP-004 badge + 5-bullet answer in bot bubble","Employee gets a concise summary — ideal for quick policy refresh","EBF3FB"),
        ],
        "result": "5-bullet credit card policy summary | CCP-004 badge | SummaryAgent chip",
    },

    "SC5 — Escalation": {
        "tab_color": "FFC00000",
        "screenshot": "sc5_escalation.png",
        "question": "I need to escalate a KYC dispute to the compliance team",
        "intent": "escalation",
        "steps": [
            ("1","Employee","UI Input","Types escalation request containing 'escalate' and 'compliance team'","Escalation triggers: escalate, contact, manager, human, speak to, urgent, compliance team","DCE6F1"),
            ("2","app.py","Streamlit","Calls orchestrator.answer(question)","Standard pipeline entry","EBF3FB"),
            ("3","pre_query_hook","Hook","Valid — passes","Logs incoming","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","Detects 'escalate' + 'compliance team' in triggers → intent = escalation","Escalation takes precedence — once detected, no policy search routing occurs","EBF3FB"),
            ("5","OrchestratorAgent","Orchestrator","Intent = escalation → routes to EscalationAgent. Passes matched policy key if available","Escalation intent always goes to EscalationAgent — not affected by cross_reference scores","DCE6F1"),
            ("6","EscalationAgent","SubAgent 5","'kyc' found in question. Reads KYC policy file via search_policy() to confirm. Looks up ESCALATION_CONTACTS[kyc_policy] → compliance team","EscalationAgent reads policy files to infer which team to contact when no key is explicitly passed","E2EFDA"),
            ("7","format_skill.generate_escalation_email()","Skill","Generates pre-filled email draft: To: compliance@bank.internal, Subject: KYC Dispute Escalation, body with employee's question","Email is generated from a template + the employee's question — never auto-sent","EBF3FB"),
            ("8","EscalationAgent returns","SubAgent 5","Returns {answer, policy_name, escalated:True, email_draft:'Dear Compliance...'} ","escalated=True flag prevents escalation_hook from firing later","DCE6F1"),
            ("9","escalation_hook","Hook","found is not False (escalated=True) — condition: found=False AND escalated=False. NOT met → hook suppressed","This is the key distinction: EscalationAgent sets escalated=True so escalation_hook does NOT fire","E2EFDA"),
            ("10","AuditLoggerAgent","SubAgent 7","Logs escalation event. comply risk=False here","Escalation logged to query_audit.log with ESCALATED status","EBF3FB"),
            ("11","app.py UI","Streamlit","Contact info + EscalationAgent chip. Email draft in expandable section '📧 View Email Draft'","The email_draft field in result dict triggers the expander in app.py — employee clicks to expand","DCE6F1"),
        ],
        "result": "Contact details + email draft expander | EscalationAgent chip | escalated=True | escalation_hook suppressed",
    },

    "SC6 — Compliance Alert": {
        "tab_color": "FFB44A00",
        "screenshot": "sc6_compliance_alert.png",
        "question": "How do I bypass KYC with a fake document?",
        "intent": "policy_query",
        "steps": [
            ("1","Employee","UI Input","Types query containing compliance risk keywords 'bypass' and 'fake document'","This demo shows the compliance detection layer. The employee may not realise this will be flagged","DCE6F1"),
            ("2","app.py","Streamlit","Calls orchestrator.answer(question)","Standard pipeline entry","EBF3FB"),
            ("3","pre_query_hook","Hook","Valid length — passes. Logs incoming query","The pre_query_hook does NOT check for compliance risk — that is AuditLoggerAgent's job","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","'bypass', 'fake' not in summary/escalation triggers. 'kyc' found → policy_query","Intent classified normally — compliance detection happens later","EBF3FB"),
            ("5","PolicySearchAgent","SubAgent 2","KYC keywords match. Finds policy content. Returns found=True","Normal policy search proceeds — the answer is about the KYC policy","DCE6F1"),
            ("6","AuditLoggerAgent.log_answer()","SubAgent 7","Calls detect_compliance_risk('How do I bypass KYC with a fake document?') → checks 14 keywords: finds 'bypass' and 'fake document' → returns True","detect_compliance_risk() checks: fraud, bypass, fake document, money laundering, sanction, bribe, and 8 others","E2EFDA"),
            ("7","compliance_alert_hook","Hook","AuditLoggerAgent compliance_risk=True → Orchestrator fires compliance_alert_hook(question, session_id). Hook logs COMPLIANCE_ALERT to session_events.log. Returns {alert: 'COMPLIANCE ALERT: Your query contains restricted keywords. This has been logged.'}","The alert string is returned and placed in result['compliance_alert'] — app.py checks this field","EBF3FB"),
            ("8","OrchestratorAgent","Orchestrator","Adds compliance_alert string to result dict. Continues returning result normally","The policy answer is still returned — the alert does not block the answer","DCE6F1"),
            ("9","post_answer_hook","Hook","Logs response time and ANSWERED status","Normal post-answer logging proceeds","E2EFDA"),
            ("10","app.py UI","Streamlit","Checks result['compliance_alert'] → not None → renders yellow warning banner ABOVE the answer. Then renders the normal answer below","The yellow banner appears at the top of the response. Employee sees both the warning AND the answer","EBF3FB"),
        ],
        "result": "Yellow compliance warning banner at top | KYC answer below | COMPLIANCE_ALERT in audit log",
    },

    "SC7 — Fallback Self-Heal": {
        "tab_color": "FFB44A00",
        "screenshot": "sc7_fallback_selfheal.png",
        "question": "What is the HR leave policy for annual leave?",
        "intent": "policy_query → fallback",
        "steps": [
            ("1","Employee","UI Input","Types question about HR leave — not covered in any of the 5 bank policies","The employee doesn't know the system only covers 5 specific policies","DCE6F1"),
            ("2","app.py","Streamlit","Calls orchestrator.answer(question)","Standard pipeline — the employee sees nothing unusual yet","EBF3FB"),
            ("3","pre_query_hook","Hook","Valid length — passes","Logs incoming","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","'leave', 'annual', 'hr' not in summary/escalation triggers → policy_query","Classified as policy_query — the search will attempt to find a match","EBF3FB"),
            ("5","PolicySearchAgent","SubAgent 2","search_policy() runs KEYWORD_MAP against all 5 policies. 'leave', 'annual', 'hr' match ZERO keywords in any policy. Returns {found: False, score: 0}","KEYWORD_MAP has no HR or leave keywords — all 5 policies score 0","DCE6F1"),
            ("6","OrchestratorAgent (self-healing)","Orchestrator","Detects found=False from PolicySearchAgent → automatically invokes FallbackAgent. No employee action required","This is the self-healing trigger. The employee never sees 'PolicySearchAgent failed' — it's invisible","E2EFDA"),
            ("7","FallbackAgent","SubAgent 6","Calls cross_reference() at low threshold (score > 0). All policies return 0 for this question. No partial match possible. Returns {answer: 'no match', needs_escalation: True, found: False}","FallbackAgent tries one more time too — but since no keywords match at all, truly nothing found","EBF3FB"),
            ("8","escalation_hook","Hook","found=False AND escalated=False → condition MET. Hook fires. Logs question to unresolved_queries.log","This is when escalation_hook fires — ONLY in this path. Note: NOT escalation intent — this is auto-escalation from a failed search","DCE6F1"),
            ("9","AuditLoggerAgent","SubAgent 7","Logs to query_audit.log with STATUS=FALLBACK. Also logs to unresolved_queries.log","Both audit logs get entries — query_audit.log marks it FALLBACK; unresolved_queries.log records it for compliance","E2EFDA"),
            ("10","post_answer_hook","Hook","Logs FALLBACK status + response time","post_answer_hook always runs — even for fallback answers","EBF3FB"),
            ("11","app.py UI","Streamlit","FallbackAgent chip shown. Amber-coloured warning bubble with contact details. No policy badge (found=False)","The amber colour signals 'no match' — distinct from the green bot bubble used for successful answers","DCE6F1"),
        ],
        "result": "Amber no-match bubble | FallbackAgent chip | escalation_hook fired | unresolved_queries.log updated",
    },

    "SC8 — Account Opening": {
        "tab_color": "FF006464",
        "screenshot": "sc8_account_opening.png",
        "question": "What is the minimum balance for a savings account?",
        "intent": "policy_query",
        "steps": [
            ("1","Employee","UI Input","Types question about savings account minimum balance","Tests the account_opening_policy keyword cluster","DCE6F1"),
            ("2","app.py","Streamlit","Standard pipeline entry","Single OrchestratorAgent handle from cache","EBF3FB"),
            ("3","pre_query_hook","Hook","Valid input — passes","Logs and sanitizes","E2EFDA"),
            ("4","QueryClassifierAgent","SubAgent 1","'account', 'savings', 'minimum', 'balance' — none in summary/escalation → policy_query","Clear policy search intent","EBF3FB"),
            ("5","OrchestratorAgent","Orchestrator","cross_reference: account_opening scores 3 ('account' + 'savings' + 'balance'). Others score 0. Single domain → PolicySearchAgent","Strong single-domain signal — no multi-policy routing","DCE6F1"),
            ("6","search_skill.search_policy()","Skill","KEYWORD_MAP account_opening_policy matches 'account', 'savings account', 'minimum balance'. Score=3. Highest across all 5. Returns snippet","'minimum balance' is a compound keyword in KEYWORD_MAP — exact phrase match scores higher","E2EFDA"),
            ("7","PolicySearchAgent","SubAgent 2","format_answer() applied. Returns {found:True, policy_key:'account_opening_policy', policy_name:'Account Opening Policy (AOP-005)', score:3}","format_answer cleans the raw snippet and adds the policy reference header","EBF3FB"),
            ("8","AuditLoggerAgent","SubAgent 7","Logs to query_audit.log. compliance_risk=False","Routine audit log","DCE6F1"),
            ("9","post_answer_hook","Hook","ANSWERED status logged","Normal post-answer audit","E2EFDA"),
            ("10","app.py UI","Streamlit","Green AOP-005 badge + PolicySearchAgent chip + answer about minimum balance requirements","Employee gets precise minimum balance figures directly from the approved policy","EBF3FB"),
        ],
        "result": "Account policy answer | AOP-005 badge | PolicySearchAgent chip | Audit logged",
    },
}


def sheet_scenario(sheet_name, data):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    tab_color = data.get("tab_color", "004472C4")
    ws.sheet_properties.tabColor = tab_color

    # Screenshot on the right
    screenshot_path = os.path.join(SCREENSHOTS, data["screenshot"])
    if os.path.exists(screenshot_path):
        img = XLImage(screenshot_path)
        img.width  = 540
        img.height = 332
        img.anchor = "M2"
        ws.add_image(img)

    # Column widths
    col_widths = [4, 22, 16, 36, 38, 22]
    headers = ["Step", "Component", "Type", "What Happens", "Talking Point (For Presenter)", "Log / Side Effect"]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(13, 30):
        ws.column_dimensions[get_column_letter(i)].width = 4.5

    # Title block
    merge_title(ws, 1, 1, 6, f"🏦  {sheet_name}  — Agent Communication Flow", tab_color, WHITE, 14, 36)
    ws.row_dimensions[1].height = 36

    # Question / intent / result info box
    info_items = [
        ("💬 User Question:", data["question"], "FFDCE6F1", "FF004472"),
        ("⚡ Intent Detected:", data["intent"], "FFE2EFDA", "FF00823E"),
        ("✅ Final Result:", data["result"], "FFFFEB9C", "FF7A3800"),
    ]
    for r_off, (label, value, bg, fg) in enumerate(info_items, start=2):
        ws.row_dimensions[r_off].height = 24
        c1 = ws.cell(row=r_off, column=1, value=label)
        c1.font = font(size=10, bold=True, color=fg)
        c1.fill = fill(bg)
        c1.alignment = align("right", "center")
        c1.border = thin_border()
        ws.merge_cells(start_row=r_off, start_column=2, end_row=r_off, end_column=6)
        c2 = ws.cell(row=r_off, column=2, value=value)
        c2.font = font(size=10, bold=False, color=fg)
        c2.fill = fill(bg)
        c2.alignment = align("left", "center", wrap=True)
        c2.border = thin_border()

    # Communication flow table header
    section_label(ws, 5, 1, 6, f"  📡  Step-by-Step Agent Communication Flow ({len(data['steps'])} steps)", tab_color)
    header_row(ws, 6, headers, col_widths, f"FF{tab_color[-6:]}" if tab_color.startswith("FF") else tab_color)

    # Add log column values
    log_map = {
        "Hook": "session_events.log",
        "SubAgent 7": "query_audit.log + session_events.log",
        "Streamlit": "—",
        "UI Input": "—",
        "Orchestrator": "—",
        "SubAgent 1": "—",
        "SubAgent 2": "—",
        "SubAgent 3": "—",
        "SubAgent 4": "—",
        "SubAgent 4 internal": "—",
        "SubAgent 5": "—",
        "SubAgent 5 returns": "—",
        "SubAgent 6": "unresolved_queries.log",
        "Skill": "—",
    }

    for r_off, step_data in enumerate(data["steps"], start=7):
        ws.row_dimensions[r_off].height = 52
        step_no, component, stype, what, talking, fll_hex = step_data
        log = log_map.get(stype, "—")
        if "hook" in component.lower():
            log = "session_events.log"
        if "escalation_hook" in component.lower() and "found=False" in what:
            log = "unresolved_queries.log"
        if "compliance_alert_hook" in component.lower():
            log = "session_events.log (COMPLIANCE_ALERT)"
        if "audit" in component.lower():
            log = "query_audit.log  +  session_events.log"

        row_vals = [step_no, component, stype, what, talking, log]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r_off, column=ci, value=val)
            c.fill = fill(f"FF{fll_hex}")
            c.border = thin_border()
            c.alignment = align("left", "top", wrap=True)
            c.font = font(size=9, bold=(ci <= 3))

    return ws


# ════════════════════════════════════════════════════════════════════════════
# SHEET 11: AGENT COMMUNICATION MAP
# ════════════════════════════════════════════════════════════════════════════
def sheet_comm_map():
    ws = wb.create_sheet("Agent Communication Map")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "004472C4"

    col_widths = [20, 18, 20, 28, 20, 20, 25]
    merge_title(ws, 1, 1, 7, "🏦  NexaBank Assistant — Complete Agent Communication Map", "001A366B", WHITE, 14, 38)

    section_label(ws, 2, 1, 7, "  📡  Who Calls Whom — Full Agent × Component Matrix", MID_BLUE)
    header_row(ws, 3,
               ["Source", "Calls", "Target", "Purpose", "Input Passed", "Output Returned", "Log Written"],
               col_widths, DARK_BLUE)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    comm_rows = [
        ("app.py (Streamlit)",     "→",  "OrchestratorAgent",       "Entry point for every question",            "question str",                      "10-field result dict",                   "—",                          LIGHT_BLUE[2:], DARK_TEXT),
        ("OrchestratorAgent",      "→",  "pre_query_hook",           "Validate and sanitize input",               "question, session_id",               "{valid, question, reason}",              "session_events.log",         "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "QueryClassifierAgent",     "Detect intent (4 types)",                   "question",                           "{intent, keywords}",                     "—",                          "DCE6F1", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "search_skill.cross_reference()", "Score all 5 policies for routing",   "question, policies",                  "list of {key, score, snippet}",          "—",                          "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "PolicySearchAgent",        "Single-domain policy query",                "question, policies",                  "{answer, policy_key, found, score}",     "—",                          "DCE6F1", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "MultiPolicyAgent",         "Cross-domain query (top≥2 & second≥1)",     "question, policies",                  "{answer, policy_names, multi, found}",   "—",                          "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "SummaryAgent",             "Policy summary request",                    "question, policies",                  "{answer, policy_name, found}",           "—",                          "DCE6F1", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "EscalationAgent",          "Explicit escalation intent",                "question, optional policy key",        "{answer, escalated, email_draft}",       "—",                          "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "FallbackAgent",            "Self-healing — auto after found=False",     "question, policies",                  "{answer, suggestion, needs_escalation}", "—",                          "DCE6F1", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "AuditLoggerAgent",         "Logs every answer (always runs last)",      "question, policy, found, session_id", "{logged, compliance_risk, session_id}",  "query_audit.log",            "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "compliance_alert_hook",    "Fires when AuditLogger flags risk",         "question, session_id",                "{alert: str}",                           "session_events.log",         "DCE6F1", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "escalation_hook",          "Fires ONLY when found=False AND NOT escalated", "question, session_id",            "None",                                   "unresolved_queries.log",     "EBF3FB", DARK_TEXT),
        ("OrchestratorAgent",      "→",  "post_answer_hook",         "Logs timing and outcome for every query",   "question, result, session_id, ms",   "None",                                   "session_events.log",         "DCE6F1", DARK_TEXT),
        ("PolicySearchAgent",      "→",  "search_skill.search_policy()", "Find best matching policy",             "question, policies",                  "{matched_key, snippet, score}",          "—",                          "EBF3FB", DARK_TEXT),
        ("PolicySearchAgent",      "→",  "format_skill.format_answer()", "Clean and format policy snippet",       "snippet, policy_key",                 "formatted str",                          "—",                          "DCE6F1", DARK_TEXT),
        ("MultiPolicyAgent",       "→",  "search_skill.cross_reference()", "Get top-2 scoring policies",          "question, policies",                  "list[{key, score, snippet}]",            "—",                          "EBF3FB", DARK_TEXT),
        ("MultiPolicyAgent",       "→",  "format_skill.format_answer()", "Format each policy snippet (×2)",       "snippet, policy_key (×2)",            "formatted str (×2)",                     "—",                          "DCE6F1", DARK_TEXT),
        ("SummaryAgent",           "→",  "SummaryAgent._detect_policy_key()", "Resolve policy via POLICY_NAME_HINTS","question",                          "policy_key str",                         "—",                          "EBF3FB", DARK_TEXT),
        ("SummaryAgent",           "→",  "search_skill.search_policy()", "Fallback: infer policy key via search",  "question, policies (fallback only)",  "{matched_key, ...}",                     "—",                          "DCE6F1", DARK_TEXT),
        ("SummaryAgent",           "→",  "format_skill.summarize_policy()", "Generate 5-bullet summary",          "policy_text, policy_key",             "5-bullet str",                           "—",                          "EBF3FB", DARK_TEXT),
        ("EscalationAgent",        "→",  "search_skill.search_policy()", "Infer policy key when not provided",    "question, policies",                  "{matched_key, ...}",                     "—",                          "DCE6F1", DARK_TEXT),
        ("EscalationAgent",        "→",  "format_skill.generate_escalation_email()", "Draft escalation email",    "question, matched_policy",            "email body str",                         "—",                          "EBF3FB", DARK_TEXT),
        ("FallbackAgent",          "→",  "search_skill.cross_reference()", "Partial match at low threshold",      "question, policies",                   "list (may be empty)",                    "—",                          "DCE6F1", DARK_TEXT),
        ("AuditLoggerAgent",       "→",  "audit_skill.log_query()",  "Write query record",                        "question, policy, found, session_id", "None",                                   "query_audit.log",            "EBF3FB", DARK_TEXT),
        ("AuditLoggerAgent",       "→",  "audit_skill.log_unresolved()", "Write unresolved record (if not found)","question, session_id",               "None",                                   "unresolved_queries.log",     "DCE6F1", DARK_TEXT),
        ("AuditLoggerAgent",       "→",  "audit_skill.detect_compliance_risk()", "Check 14 risk keywords",        "question",                            "bool",                                   "—",                          "EBF3FB", DARK_TEXT),
        ("QueryClassifierAgent",   "→",  "search_skill.extract_keywords()", "Extract meaningful words",           "question",                            "list[str]",                              "—",                          "DCE6F1", DARK_TEXT),
        ("app.py",                 "→",  "session_start_hook",       "On app first load",                         "session_id",                          "None",                                   "session_events.log",         "EBF3FB", DARK_TEXT),
        ("app.py (Clear Chat)",    "→",  "session_end_hook",         "On Clear Chat button click",               "session_id, total_queries",            "None",                                   "session_events.log",         "DCE6F1", DARK_TEXT),
    ]

    for r_off, row in enumerate(comm_rows, start=4):
        ws.row_dimensions[r_off].height = 36
        src, arrow, tgt, purpose, inp, out, log, fll, _ = row
        for ci, val in enumerate([src, arrow, tgt, purpose, inp, out, log], start=1):
            c = ws.cell(row=r_off, column=ci, value=val)
            c.fill = fill(f"FF{fll}")
            c.border = thin_border()
            c.alignment = align("left", "top", wrap=True)
            c.font = font(size=9, bold=(ci in [1,3]),
                         color=("FF004472" if ci==3 else "FF1A1A2E"))

    return ws


# ════════════════════════════════════════════════════════════════════════════
# SHEET 12: SKILL & HOOK REFERENCE
# ════════════════════════════════════════════════════════════════════════════
def sheet_skill_hook_reference():
    ws = wb.create_sheet("Skill & Hook Reference")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0000823E"

    col_widths = [5, 22, 18, 22, 22, 26, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 7, "🏦  NexaBank Assistant — Skill & Hook Complete Reference", "00006400", WHITE, 14, 38)

    # SKILLS
    section_label(ws, 2, 1, 7, "  🔧  Skills — 3 Modules  (skills/)", "00006400")
    header_row(ws, 3, ["#","File","Function","Parameters","Returns","Purpose","Called By"], col_widths, "00006400")
    skills = [
        ("1","search_skill.py","search_policy()","question, policies","{ matched_key, policy_text, snippet, score }","Best-matching policy via KEYWORD_MAP scoring across all 5 policies","PolicySearchAgent, SummaryAgent (fallback), EscalationAgent(infer)","E2EFDA"),
        ("2","search_skill.py","extract_keywords()","text","list[str]","Extracts words > 3 chars; removes stopwords. Used to classify intent","QueryClassifierAgent","E2EFDA"),
        ("3","search_skill.py","cross_reference()","question, policies","list[{matched_key, score, snippet}] sorted desc","Scores ALL policies — drives multi-policy routing and FallbackAgent partial recovery","OrchestratorAgent (routing), MultiPolicyAgent, FallbackAgent","E2EFDA"),
        ("4","format_skill.py","POLICY_DISPLAY_NAMES","—","dict","Canonical map of policy keys → display names. Single source of truth","All agents (import)","DCE6F1"),
        ("5","format_skill.py","format_answer()","raw_snippet, policy_key","str","Cleans raw policy extract; adds header with policy name and reference code","PolicySearchAgent, MultiPolicyAgent","DCE6F1"),
        ("6","format_skill.py","summarize_policy()","policy_text, policy_key","5-bullet str","Generates human-readable policy summary from full policy text","SummaryAgent","DCE6F1"),
        ("7","format_skill.py","generate_escalation_email()","question, matched_policy","email body str","Drafts pre-filled escalation email with correct contact and employee's question","EscalationAgent","DCE6F1"),
        ("8","audit_skill.py","log_query()","question, policy, found, session_id","None","Appends timestamped record to query_audit.log with MATCHED or FALLBACK status","AuditLoggerAgent","FFFEB9"),
        ("9","audit_skill.py","log_unresolved()","question, session_id","None","Appends unmatched question to unresolved_queries.log for compliance review","AuditLoggerAgent, escalation_hook","FFFEB9"),
        ("10","audit_skill.py","log_session_event()","event, session_id, detail","None","Appends lifecycle event to session_events.log (start, end, escalation, alert)","All 6 hooks","FFFEB9"),
        ("11","audit_skill.py","detect_compliance_risk()","question","bool","Returns True if any of 14 risk keywords found: fraud, bypass, fake document, money laundering, sanction, bribe…","AuditLoggerAgent","FFFEB9"),
        ("12","audit_skill.py","read_audit_log()","log_name, last_n","list[str]","Reads last N lines from a log file — used by app.py admin audit panel","app.py Audit Log tab","FFFEB9"),
    ]
    for r_off, row in enumerate(skills, start=4):
        ws.row_dimensions[r_off].height = 42
        *vals, fll = row
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r_off, column=ci, value=v)
            c.fill = fill(f"FF{fll}")
            c.border = thin_border()
            c.alignment = align("left", "top", wrap=True)
            c.font = font(size=9, bold=(ci<=2))

    # HOOKS
    row_start = 4 + len(skills) + 1
    section_label(ws, row_start, 1, 7, "  🪝  Hooks — 6 Lifecycle Callbacks  (hooks.py)", "004472C4")
    header_row(ws, row_start+1,
               ["#","Hook Function","Trigger Condition","Inputs","Returns","Side Effect / Log","Key Behaviour"],
               col_widths, MID_BLUE)
    hooks = [
        ("1","pre_query_hook","BEFORE any agent — every query","question, session_id","{ valid: bool, question: str, reason: str }","session_events.log — logs INCOMING_QUERY","Rejects empty or < 3-char questions. First gate in pipeline","DCE6F1"),
        ("2","post_answer_hook","AFTER full pipeline — every query","question, result, session_id, response_ms","None","session_events.log — logs ANSWERED or FALLBACK + response time","Creates permanent timing + outcome record for every interaction","E2EFDA"),
        ("3","session_start_hook","OrchestratorAgent.__init__() — once per session","session_id","None","session_events.log — logs SESSION_START","Fires once when employee first loads the app. Marks session start with timestamp","DCE6F1"),
        ("4","session_end_hook","Employee clicks Clear Chat button","session_id, total_queries","None","session_events.log — logs SESSION_END + total_queries count","Marks clean session close; total query count recorded for usage analytics","E2EFDA"),
        ("5","escalation_hook","found=False AND escalated=False ONLY","question, session_id","None","unresolved_queries.log — logs UNRESOLVED_QUERY","Does NOT fire when EscalationAgent ran (escalated=True). Only fires on genuine no-match scenarios","DCE6F1"),
        ("6","compliance_alert_hook","detect_compliance_risk() = True (14 risk keywords)","question, session_id","{ alert: str }","session_events.log — logs COMPLIANCE_ALERT","Returns alert string placed in result dict. app.py renders yellow warning banner when this is non-null","E2EFDA"),
    ]
    for r_off, row in enumerate(hooks, start=row_start+2):
        ws.row_dimensions[r_off].height = 52
        *vals, fll = row
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r_off, column=ci, value=v)
            c.fill = fill(f"FF{fll}")
            c.border = thin_border()
            c.alignment = align("left", "top", wrap=True)
            c.font = font(size=9, bold=(ci<=2))

    return ws


# ════════════════════════════════════════════════════════════════════════════
# BUILD EVERYTHING
# ════════════════════════════════════════════════════════════════════════════
print("Building Excel user guide...")

print("  → Cover sheet")
sheet_cover()

print("  → Scenario Overview")
sheet_scenario_overview()

for name, data in SCENARIO_DATA.items():
    print(f"  → {name}")
    sheet_scenario(name, data)

print("  → Agent Communication Map")
sheet_comm_map()

print("  → Skill & Hook Reference")
sheet_skill_hook_reference()

wb.save(OUT)
print(f"\n✅  Saved: {OUT}")
print(f"    Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"      • {s}")
